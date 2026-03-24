import sys
sys.stdout.reconfigure(encoding='utf-8')

# ============================================================
# JUNE 2026 SCHEDULE COMPARISON: AI (CP-SAT) vs HUMAN
# ============================================================

import openpyxl
from collections import defaultdict
import math
import yaml


def parse_all_slots(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    all_slots = []
    day_cols = {}
    i = 0
    while i < len(rows):
        row = rows[i]
        v1 = str(row[1]).strip() if row[1] is not None else ''
        if v1 in ('SUN', 'MON'):
            day_row = rows[i+1] if i+1 < len(rows) else row
            day_cols = {}
            for ci, val in enumerate(day_row):
                if val is None:
                    continue
                try:
                    day_cols[ci] = int(float(str(val)))
                except Exception:
                    pass
            i += 2
            continue
        v0 = str(row[0]).strip() if row[0] is not None else ''
        is_time = (len(v0) >= 4 and '-' in v0 and any(c.isdigit() for c in v0[:4]))
        if v0 and not is_time and v0 not in ('None', ''):
            if i+1 < len(rows):
                next_v0 = str(rows[i+1][0]).strip() if rows[i+1][0] is not None else ''
                next_is_time = (len(next_v0) >= 4 and '-' in next_v0 and any(c.isdigit() for c in next_v0[:4]))
                if next_is_time:
                    for ci in day_cols:
                        physician = row[ci] if ci < len(row) else None
                        p = str(physician).strip() if physician is not None else None
                        if p in ('None', '', '---', None):
                            p = None
                        all_slots.append({
                            'date': day_cols[ci],
                            'slot': v0.strip(),
                            'physician': p,
                            'shift_time': next_v0
                        })
                    i += 2
                    continue
        i += 1
    return all_slots


def load_requested(path):
    wb = openpyxl.load_workbook(path)
    ws = wb['Preferences']
    rows = list(ws.iter_rows(values_only=True))
    requested = {}
    for row in rows[1:]:
        if row[0] and row[1] and '2026-06' in str(row[1]):
            physician = str(row[0]).strip()
            if physician not in requested and row[4] is not None:
                requested[physician] = int(row[4])
    return requested


def classify_group(shift_time, slot):
    if slot in ('DOC', 'Day On Call'):
        return 'A'
    if slot in ('NOC', 'Night On Call'):
        return 'B'
    st = shift_time.replace(' ', '') if shift_time else ''
    try:
        start = int(st.split('-')[0][:4])
    except Exception:
        return 'A'
    if start >= 1600:
        return 'B'
    return 'A'


def find_singletons(assignments):
    doc_day = defaultdict(lambda: defaultdict(list))
    for a in assignments:
        doc_day[a['physician']][a['date']].append(a)
    singletons = []
    for doc, days in doc_day.items():
        sorted_dates = sorted(days.keys())
        for i, d in enumerate(sorted_dates):
            prev_d = sorted_dates[i-1] if i > 0 else None
            next_d = sorted_dates[i+1] if i+1 < len(sorted_dates) else None
            has_prev = (prev_d == d - 1)
            has_next = (next_d == d + 1)
            if not has_prev and not has_next:
                singletons.append({'physician': doc, 'date': d})
    return singletons


def find_consecutive_violations(assignments, physician_max_map):
    doc_day = defaultdict(set)
    for a in assignments:
        doc_day[a['physician']].add(a['date'])
    violations = []
    for doc, days in doc_day.items():
        sorted_dates = sorted(days)
        max_consec = physician_max_map.get(doc, 3)
        run_len = 1
        run_start = sorted_dates[0]
        for i in range(1, len(sorted_dates)):
            if sorted_dates[i] == sorted_dates[i-1] + 1:
                run_len += 1
            else:
                if run_len > max_consec:
                    violations.append({
                        'physician': doc, 'start': run_start,
                        'length': run_len, 'max': max_consec
                    })
                run_len = 1
                run_start = sorted_dates[i]
        if run_len > max_consec:
            violations.append({
                'physician': doc, 'start': run_start,
                'length': run_len, 'max': max_consec
            })
    return violations


# Load data
ai_all = parse_all_slots(r'C:/Users/kskob/Dropbox/KEAclaude/KEAsked/Request-Imports/cpsat-schedule.xlsx')
hu_all = parse_all_slots(r'C:/Users/kskob/Dropbox/KEAclaude/KEAsked/Request-Imports/June26-complete.xlsx')
ai_raw = [a for a in ai_all if a['physician']]
hu_raw = [a for a in hu_all if a['physician']]

name_map_hu = {
    'Lam N': 'N Lam', 'Hanson A': 'Amanda Hanson', 'Chang J': 'J Chang',
    'MacKenzie': 'macKenzie', 'Yeung, Aref': 'Aref Yeung', 'Scheirer R': 'R Scheirer',
    'Dong S': 'Dong', 'Yeung, Alex': 'Alex Yeung', 'Chang E': 'E Chang',
    'Braun': 'BRAUN', 'Brown F': 'FBrown', 'Brown T': 'TBrown',
    'Scheirer O': 'O Scheirer', 'Peterson C': 'Peterson', 'Zhang': 'ZHANG',
}

for a in hu_raw:
    a['physician'] = name_map_hu.get(a['physician'].strip(), a['physician'].strip())

requested_shifts = load_requested(
    r'C:/Users/kskob/Dropbox/KEAclaude/KEAsked/Request-Imports/SingleJune.xlsx'
)

for a in ai_raw:
    a['group'] = classify_group(a['shift_time'], a['slot'])
for a in hu_raw:
    a['group'] = classify_group(a['shift_time'], a['slot'])

with open(r'C:/Users/kskob/Dropbox/KEAclaude/KEAsked/scheduler/config/physicians.yaml', 'r') as f:
    yaml_data = yaml.safe_load(f)
physician_max_yaml = {}
for p in yaml_data['physicians']:
    physician_max_yaml[p['name']] = p['scheduling'].get('max_consecutive_shifts', 3)

yaml_name_map = {
    'Amanda Hanson': 'Amanda Hanson', 'Aref Yeung': 'Aref Yeung',
    'BRAUN': 'Braun', 'Alex Yeung': 'Yeung Alex', 'ZHANG': 'Zhang',
    'macKenzie': 'Mackenzie', 'N Lam': 'Lam N', 'Lam-Rico': 'Lam-Rico',
    'FBrown': 'Brown F', 'TBrown': 'Brown T', 'R Scheirer': 'R Scheirer',
    'O Scheirer': 'O Scheirer', 'E Chang': 'E Chang', 'J Chang': 'J Chang',
    'Skoblenick': 'Kevin Skoblenick',
}


def get_max_consec(doc):
    if doc in physician_max_yaml:
        return physician_max_yaml[doc]
    mapped = yaml_name_map.get(doc, doc)
    return physician_max_yaml.get(mapped, 3)


ai_doc_max = {doc: get_max_consec(doc) for doc in set(a['physician'] for a in ai_raw)}
hu_doc_max = {doc: get_max_consec(doc) for doc in set(a['physician'] for a in hu_raw)}

ai_counts = defaultdict(int)
for a in ai_raw:
    ai_counts[a['physician']] += 1
hu_counts = defaultdict(int)
for a in hu_raw:
    hu_counts[a['physician']] += 1

ai_total = len(ai_all)
ai_filled_n = len(ai_raw)
ai_unfilled_n = ai_total - ai_filled_n
hu_total = len(hu_all)
hu_filled_n = len(hu_raw)
hu_unfilled_n = hu_total - hu_filled_n
ai_unfilled_slots = [(a['date'], a['slot'], a['shift_time']) for a in ai_all if not a['physician']]
hu_unfilled_slots = [(a['date'], a['slot'], a['shift_time']) for a in hu_all if not a['physician']]

all_physicians = sorted(set(list(ai_counts.keys()) + list(hu_counts.keys()) + list(requested_shifts.keys())))
discrepancies = []
for doc in all_physicians:
    req = requested_shifts.get(doc, None)
    hc = hu_counts.get(doc, 0)
    ac = ai_counts.get(doc, 0)
    h_diff = hc - req if req is not None else None
    a_diff = ac - req if req is not None else None
    discrepancies.append((doc, req, hc, ac, h_diff, a_diff))

discrepancies_sorted = sorted(
    discrepancies, key=lambda x: abs(x[5]) if x[5] is not None else 0, reverse=True
)
n_with_req = sum(1 for d, r, h, a, hd, ad in discrepancies if r is not None)
ai_at_req = sum(1 for d, r, h, a, hd, ad in discrepancies if r is not None and ad == 0)
hu_at_req = sum(1 for d, r, h, a, hd, ad in discrepancies if r is not None and hd == 0)
ai_rmse_shifts = math.sqrt(
    sum(ad**2 for _, r, h, a, hd, ad in discrepancies if r is not None and ad is not None) / n_with_req
)
hu_rmse_shifts = math.sqrt(
    sum(hd**2 for _, r, h, a, hd, ad in discrepancies if r is not None and hd is not None) / n_with_req
)


def ab_breakdown(assignments):
    counts = defaultdict(lambda: {'A': 0, 'B': 0})
    for a in assignments:
        counts[a['physician']][a['group']] += 1
    return counts


ai_ab = ab_breakdown(ai_raw)
hu_ab = ab_breakdown(hu_raw)
all_docs = sorted(set(list(ai_ab.keys()) + list(hu_ab.keys())))

ab_data = []
for doc in all_docs:
    ha = hu_ab[doc]['A']
    hb = hu_ab[doc]['B']
    aa = ai_ab[doc]['A']
    ab2 = ai_ab[doc]['B']
    h_pct = ha / (ha+hb) * 100 if (ha+hb) > 0 else None
    a_pct = aa / (aa+ab2) * 100 if (aa+ab2) > 0 else None
    ai_gap = abs(a_pct - 40) if a_pct is not None else 999
    hu_gap = abs(h_pct - 40) if h_pct is not None else 999
    ab_data.append((doc, h_pct, a_pct, ai_gap, hu_gap))

ab_data_sorted = sorted(ab_data, key=lambda x: x[3], reverse=True)
valid_ai_ab = [x for x in ab_data if x[3] < 999]
valid_hu_ab = [x for x in ab_data if x[4] < 999]
ai_rmse_ab = math.sqrt(sum(x[3]**2 for x in valid_ai_ab) / len(valid_ai_ab))
hu_rmse_ab = math.sqrt(sum(x[4]**2 for x in valid_hu_ab) / len(valid_hu_ab))
ai_allA = sum(1 for d, hp, ap, aig, hug in ab_data if ap is not None and ap == 100)
ai_allB = sum(1 for d, hp, ap, aig, hug in ab_data if ap is not None and ap == 0)
hu_allA = sum(1 for d, hp, ap, aig, hug in ab_data if hp is not None and hp == 100)
hu_allB = sum(1 for d, hp, ap, aig, hug in ab_data if hp is not None and hp == 0)

ai_singletons = find_singletons(ai_raw)
hu_singletons = find_singletons(hu_raw)
ai_sing_counts = defaultdict(int)
for s in ai_singletons:
    ai_sing_counts[s['physician']] += 1
hu_sing_counts = defaultdict(int)
for s in hu_singletons:
    hu_sing_counts[s['physician']] += 1

ai_violations = find_consecutive_violations(ai_raw, ai_doc_max)
hu_violations = find_consecutive_violations(hu_raw, hu_doc_max)

SEP = "=" * 72

# ─── SECTION 1 ─────────────────────────────────────────────────────────────

# Use the intersection of slot types so both schedules are measured on identical
# shift categories — this excludes AM/PM CALL (human-only) and any slot types
# present in the AI Excel but absent from the human schedule, giving a fair denominator.
_ai_slot_types = {a['slot'] for a in ai_all}
_hu_slot_types = {a['slot'] for a in hu_all}
_common_slot_types = _ai_slot_types & _hu_slot_types
_ai_only_slots = _ai_slot_types - _hu_slot_types
_hu_only_slots = _hu_slot_types - _ai_slot_types
ai_common = [a for a in ai_all if a['slot'] in _common_slot_types]
hu_common = [a for a in hu_all if a['slot'] in _common_slot_types]
ai_com_fill = len([a for a in ai_common if a['physician']])
hu_com_fill = len([a for a in hu_common if a['physician']])

print(SEP)
print("JUNE 2026 SCHEDULE COMPARISON: AI (CP-SAT) vs HUMAN (GOLD STANDARD)")
print(SEP)
print()
print(SEP)
print("SECTION 1: FILL RATE")
print(SEP)
if _ai_only_slots:
    print(f"NOTE: AI-only slot types (excluded from common comparison): {sorted(_ai_only_slots)}")
if _hu_only_slots:
    print(f"NOTE: Human-only slot types (excluded from common comparison): {sorted(_hu_only_slots)}")
print()
print(f"{'Metric':<40} {'AI':>12} {'Human':>12}")
print("-" * 66)
print(f"{'Total defined slots (raw)':<40} {ai_total:>12} {hu_total:>12}")
print(f"{'Common slot types (shared base)':<40} {len(ai_common):>12} {len(hu_common):>12}")
print(f"{'Filled (common slots)':<40} {ai_com_fill:>12} {hu_com_fill:>12}")
print(f"{'Unfilled (common slots)':<40} {len(ai_common)-ai_com_fill:>12} {len(hu_common)-hu_com_fill:>12}")
ai_cfr = str(round(ai_com_fill/len(ai_common)*100, 1)) + '%'
hu_cfr = str(round(hu_com_fill/len(hu_common)*100, 1)) + '%'
print(f"{'Fill rate (common slots)':<40} {ai_cfr:>12} {hu_cfr:>12}")
ai_unfilled_common = [(a['date'], a['slot'], a['shift_time']) for a in ai_common if not a['physician']]
hu_unfilled_common = [(a['date'], a['slot'], a['shift_time']) for a in hu_common if not a['physician']]
print(f"\nAI unfilled common slots ({len(ai_unfilled_common)}):")
for u in sorted(ai_unfilled_common, key=lambda x: x[0]):
    print(f"  Day {u[0]:2d}  {u[1]:<15}  {u[2]}")
print(f"\nHuman unfilled common slots ({len(hu_unfilled_common)}):")
for u in sorted(hu_unfilled_common, key=lambda x: x[0]):
    print(f"  Day {u[0]:2d}  {u[1]:<15}  {u[2]}")

# ─── SECTION 2 ─────────────────────────────────────────────────────────────

print()
print(SEP)
print("SECTION 2: PER-PHYSICIAN SHIFT COUNT")
print(SEP)
print(f"{'Physician':<22} {'Req':>4} {'Human':>6} {'AI':>4} {'H-Req':>6} {'AI-Req':>7}")
print("-" * 55)
for doc, req, hc, ac, hd, ad in discrepancies_sorted:
    if req is None:
        continue
    hd_s = ('+' + str(hd) if hd >= 0 else str(hd)) if hd is not None else 'N/A'
    ad_s = ('+' + str(ad) if ad >= 0 else str(ad)) if ad is not None else 'N/A'
    flag = '  ***' if ad is not None and abs(ad) >= 3 else ''
    print(f"{doc:<22} {req:>4} {hc:>6} {ac:>4} {hd_s:>6} {ad_s:>7}{flag}")
print()
print(f"Physicians at exact requested count:  AI={ai_at_req}/{n_with_req}  Human={hu_at_req}/{n_with_req}")
print(f"Shift count RMSE:  AI={ai_rmse_shifts:.2f}  Human={hu_rmse_shifts:.2f}")

# ─── SECTION 3 ─────────────────────────────────────────────────────────────

print()
print(SEP)
print("SECTION 3: GROUP A/B BALANCE (Target 40%A / 60%B)")
print(SEP)
print("Group A = day shifts (start < 1600), Group B = evening/night (start >= 1600)")
print()
print(f"{'Physician':<22} {'HU %A':>7} {'AI %A':>7} {'HU gap':>7} {'AI gap':>7}")
print("-" * 57)
for doc, hp, ap, ai_gap, hu_gap in ab_data_sorted:
    hp_s = (str(round(hp, 1)) + '%') if hp is not None else 'N/A'
    ap_s = (str(round(ap, 1)) + '%') if ap is not None else 'N/A'
    agu = (str(round(ai_gap, 1)) + '%') if ai_gap < 999 else 'N/A'
    hgu = (str(round(hu_gap, 1)) + '%') if hu_gap < 999 else 'N/A'
    flag = '  <--' if ai_gap > 40 else ''
    print(f"{doc:<22} {hp_s:>7} {ap_s:>7} {hgu:>7} {agu:>7}{flag}")
print()
print(f"A/B RMSE from 40%:  AI={ai_rmse_ab:.2f}%  Human={hu_rmse_ab:.2f}%")
print(f"100% Group A (day-only):  AI={ai_allA}  Human={hu_allA}")
print(f"100% Group B (night-only):  AI={ai_allB}  Human={hu_allB}")

# ─── SECTION 4 ─────────────────────────────────────────────────────────────

print()
print(SEP)
print("SECTION 4: SINGLETON ANALYSIS")
print(SEP)
print(f"Total singletons:  AI={len(ai_singletons)}  Human={len(hu_singletons)}")
all_sing_docs = sorted(
    set(list(ai_sing_counts.keys()) + list(hu_sing_counts.keys())),
    key=lambda d: ai_sing_counts.get(d, 0) - hu_sing_counts.get(d, 0),
    reverse=True
)
print(f"\n{'Physician':<22} {'Human':>6} {'AI':>4} {'Diff':>6}")
print("-" * 42)
for doc in all_sing_docs:
    ac = ai_sing_counts.get(doc, 0)
    hc = hu_sing_counts.get(doc, 0)
    if ac > 0 or hc > 0:
        diff_s = ('+' + str(ac-hc) if ac-hc >= 0 else str(ac-hc))
        print(f"{doc:<22} {hc:>6} {ac:>4} {diff_s:>6}")

# ─── SECTION 5 ─────────────────────────────────────────────────────────────

print()
print(SEP)
print("SECTION 5: CONSECUTIVE VIOLATIONS")
print(SEP)
print(f"AI violations: {len(ai_violations)}")
for v in sorted(ai_violations, key=lambda x: x['length'], reverse=True):
    print(f"  {v['physician']:<22}  run={v['length']} days (max={v['max']}) start=day {v['start']}")
print(f"\nHuman violations: {len(hu_violations)}")
for v in sorted(hu_violations, key=lambda x: x['length'], reverse=True):
    print(f"  {v['physician']:<22}  run={v['length']} days (max={v['max']}) start=day {v['start']}")

# ─── SECTION 6 ─────────────────────────────────────────────────────────────

print()
print(SEP)
print("SECTION 6: KEY STRUCTURAL DIFFERENCES")
print(SEP)

ai_late_unfilled = [a for a in ai_all if a['date'] >= 20 and not a['physician']]
hu_late_unfilled = [a for a in hu_all if a['date'] >= 20 and not a['physician']]
ai_late_all = [a for a in ai_raw if a['date'] >= 20]
hu_late_all = [a for a in hu_raw if a['date'] >= 20]
print(f"Late-month (days 20-30) coverage:")
print(f"  AI:    {len(ai_late_all)} filled, {len(ai_late_unfilled)} unfilled slots")
print(f"  Human: {len(hu_late_all)} filled, {len(hu_late_unfilled)} unfilled slots")

ai_2400 = [a for a in ai_raw if '2400' in str(a['shift_time'])]
hu_2400 = [a for a in hu_raw if '2400' in str(a['shift_time'])]
print(f"\n2400h shift totals:  AI={len(ai_2400)}  Human={len(hu_2400)}")

ai_2400_by_doc = defaultdict(int)
for a in ai_2400:
    ai_2400_by_doc[a['physician']] += 1
hu_2400_by_doc = defaultdict(int)
for a in hu_2400:
    hu_2400_by_doc[a['physician']] += 1

top_2400 = sorted(
    set(list(ai_2400_by_doc.keys()) + list(hu_2400_by_doc.keys())),
    key=lambda d: ai_2400_by_doc.get(d, 0) + hu_2400_by_doc.get(d, 0),
    reverse=True
)[:15]
print(f"\n{'Physician':<22} {'AI 2400':>8} {'HU 2400':>9}")
print("-" * 42)
for doc in top_2400:
    print(f"{doc:<22} {ai_2400_by_doc.get(doc,0):>8} {hu_2400_by_doc.get(doc,0):>9}")

ai_end_docs = set(a['physician'] for a in ai_raw if a['date'] >= 25)
hu_end_docs = set(a['physician'] for a in hu_raw if a['date'] >= 25)
missing_late = hu_end_docs - ai_end_docs
print(f"\nPhysicians in Human days 25-30 but MISSING from AI days 25-30:")
for doc in sorted(missing_late):
    print(f"  {doc}")

print("\n2400h night clustering:")
for label, raw2 in [("AI", ai_raw), ("Human", hu_raw)]:
    night_by_doc = defaultdict(list)
    for a in raw2:
        if '2400' in str(a['shift_time']):
            night_by_doc[a['physician']].append(a['date'])
    clustered = 0
    isolated = 0
    for doc2, dates2 in night_by_doc.items():
        sorted_d = sorted(dates2)
        for i, d in enumerate(sorted_d):
            prev2 = sorted_d[i-1] if i > 0 else None
            nxt2 = sorted_d[i+1] if i+1 < len(sorted_d) else None
            if prev2 == d-1 or nxt2 == d+1:
                clustered += 1
            else:
                isolated += 1
    total = clustered + isolated
    if total > 0:
        print(f"  {label}: {total} 2400h slots -- {clustered} clustered ({round(clustered/total*100)}%), "
              f"{isolated} isolated ({round(isolated/total*100)}%)")

# ─── SECTION 7 ─────────────────────────────────────────────────────────────

print()
print(SEP)
print("SECTION 7: SUMMARY TABLE")
print(SEP)
print(f"{'Metric':<45} {'AI':>12} {'Human':>12}")
print("-" * 72)
print(f"{'Total defined slots':<45} {ai_total:>12} {hu_total:>12}")
print(f"{'Filled slots':<45} {ai_filled_n:>12} {hu_filled_n:>12}")
print(f"{'Unfilled slots':<45} {ai_unfilled_n:>12} {hu_unfilled_n:>12}")
print(f"{'Common-slot fill rate':<45} {ai_cfr:>12} {hu_cfr:>12}")
print(f"{'Total singletons':<45} {len(ai_singletons):>12} {len(hu_singletons):>12}")
at_ai = str(ai_at_req) + '/' + str(n_with_req)
at_hu = str(hu_at_req) + '/' + str(n_with_req)
print(f"{'Physicians at exact requested count':<45} {at_ai:>12} {at_hu:>12}")
print(f"{'Shift count RMSE':<45} {round(ai_rmse_shifts,2):>12} {round(hu_rmse_shifts,2):>12}")
ai_ab_s = str(round(ai_rmse_ab, 2)) + '%'
hu_ab_s = str(round(hu_rmse_ab, 2)) + '%'
print(f"{'A/B balance RMSE from 40%':<45} {ai_ab_s:>12} {hu_ab_s:>12}")
print(f"{'Physicians 100% Group A (day-only)':<45} {ai_allA:>12} {hu_allA:>12}")
print(f"{'Physicians 100% Group B (night-only)':<45} {ai_allB:>12} {hu_allB:>12}")
print(f"{'Consecutive violations':<45} {len(ai_violations):>12} {len(hu_violations):>12}")
print(f"{'Late-month (20-30) unfilled slots':<45} {len(ai_late_unfilled):>12} {len(hu_late_unfilled):>12}")

# ─── SECTION 8 ─────────────────────────────────────────────────────────────

print()
print(SEP)
print("SECTION 8: WHAT THE AI SHOULD LEARN FROM THE HUMAN SCHEDULE")
print(SEP)

lam_ai = [a for a in ai_raw if a['physician'] == 'Lam-Rico']
lam_hu = [a for a in hu_raw if a['physician'] == 'Lam-Rico']
lam_ai_dates = sorted(set(a['date'] for a in lam_ai))
lam_hu_dates = sorted(set(a['date'] for a in lam_hu))
print(f"1. Lam-Rico under-scheduling: Req=16, AI={len(lam_ai)} (-12!), Human={len(lam_hu)} (+3)")
print(f"   AI dates:    {lam_ai_dates}")
print(f"   Human dates: {lam_hu_dates}")

print("\n2. Physicians with ZERO AI assignments but assigned by Human:")
for doc in sorted(hu_counts.keys()):
    if ai_counts.get(doc, 0) == 0 and hu_counts[doc] > 0:
        print(f"   {doc:<22} Human={hu_counts[doc]}")

print("\n3. Physicians over-assigned by AI vs requested (AI-Req >= +3):")
for doc, req, hc, ac, hd, ad in discrepancies_sorted:
    if req is not None and ad is not None and ad >= 3:
        print(f"   {doc:<22} Req={req} AI={ac} Human={hc} (over by {ad})")

print("\n4. Physicians under-assigned by AI vs requested (AI-Req <= -3):")
for doc, req, hc, ac, hd, ad in discrepancies_sorted:
    if req is not None and ad is not None and ad <= -3:
        print(f"   {doc:<22} Req={req} AI={ac} Human={hc} (under by {abs(ad)})")

print("\n5. Day-only physicians in AI (should have ~40% night shifts):")
day_only_big = [
    (doc, hp, ap, aig, hug) for doc, hp, ap, aig, hug in ab_data
    if ap is not None and ap == 100 and (ai_ab[doc]['A'] + ai_ab[doc]['B']) >= 5
]
for doc, hp, ap, aig, hug in sorted(day_only_big, key=lambda x: ai_ab[x[0]]['A'], reverse=True)[:12]:
    total = ai_ab[doc]['A'] + ai_ab[doc]['B']
    hp_s = str(round(hp, 0)) + '%' if hp is not None else 'N/A'
    print(f"   {doc:<22} {total} shifts, 100% day -- Human had {hp_s} day")

print()
print(SEP)
print("ANALYSIS COMPLETE")
print(SEP)

TIME_MAP = {
    '0600-1200': '0600h',
    '0600-1400': '0600h',
    '0900-1700': '0900h',
    '1000-1800': '1000h',
    '1200-1800': '1200h',
    '1200-2000': '1200h',
    '1400-2200': '1400h',
    '1500-2300': '1500h',
    '1600-0459': '1600h',
    '1600-2400': '1600h',
    '1700-0100': '1700h',
    '  1800-0000': '1800h',
    '1800-0000': '1800h',
    '1800-0200': '1800h',
    '2000-0400': '2000h',
    '2400-0600': '2400h',
    '2400-0800': '2400h',
}

SITE_MAP = {
    'RAH A': 'RAH A side',
    'RAH B': 'RAH B side',
    'NECHC': 'NEHC',
    'NECHC ': 'NEHC',
    'RAH I': 'RAH I side',
    'RAH Float': 'RAH F side',
}

SKIP_SITES = {'AM CALL', 'PM CALL'}

DAY_COLS = ['SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']


def parse_schedule(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    assignments = []  # (physician, date_obj, shift_code)

    i = 0
    while i < len(rows):
        row = rows[i]

        # Detect header row: col B (index 1) == 'SUN'
        if row[1] == 'SUN':
            # Next row has date numbers
            if i + 1 >= len(rows):
                i += 1
                continue
            date_row = rows[i + 1]
            # Cols B-H (indices 1-7) have date numbers
            # We need to figure out the month/year from context
            # Dates are just day numbers; we'll collect them and resolve later
            day_numbers = []
            for col_idx in range(1, 8):  # B through H
                val = date_row[col_idx]
                if val is not None:
                    try:
                        day_numbers.append(int(val))
                    except (ValueError, TypeError):
                        day_numbers.append(None)
                else:
                    day_numbers.append(None)

            # Now read site/time pairs
            j = i + 2
            current_week_data = []  # list of (site, time_code, col_idx 1-7, physician)

            while j < len(rows):
                site_row = rows[j]
                # Check if this is another header row
                if site_row[1] == 'SUN':
                    break
                # Check if site_row col A has a site label
                site_label = str(site_row[0]).strip() if site_row[0] is not None else ''
                if site_label == '' or site_label == 'None':
                    j += 1
                    continue

                # Could be a site row - check if next row has a time
                if j + 1 < len(rows):
                    time_row = rows[j + 1]
                    # Check col A of time row for a time string
                    time_val = str(time_row[0]).strip() if time_row[0] is not None else ''

                    # Is it a time range?
                    is_time = time_val in TIME_MAP or time_val.strip() in TIME_MAP
                    if not is_time:
                        # Maybe col A is None and time is in col B?
                        # Try checking if site_label looks like a time
                        j += 1
                        continue

                    if site_label in SKIP_SITES:
                        j += 2
                        continue

                    site_code = SITE_MAP.get(site_label)
                    if site_code is None:
                        j += 2
                        continue

                    time_key = time_val if time_val in TIME_MAP else time_val.strip()
                    time_code = TIME_MAP.get(time_key, time_key)
                    shift_code = f"{time_code} {site_code}"

                    # Extract physician names from site_row cols B-H (indices 1-7)
                    for col_idx in range(1, 8):
                        physician = site_row[col_idx]
                        if physician is not None:
                            physician = str(physician).strip()
                            if physician and physician != 'None' and physician != '':
                                current_week_data.append((col_idx - 1, physician, shift_code))
                                # col_idx-1 maps to day_numbers index (0=SUN...6=SAT)

                    j += 2
                else:
                    j += 1

            # Now we need to resolve day numbers to actual dates
            # We'll store them with day_numbers and resolve in a second pass
            for (day_offset, physician, shift_code) in current_week_data:
                day_num = day_numbers[day_offset] if day_offset < len(day_numbers) else None
                if day_num is not None:
                    assignments.append((physician, day_num, shift_code, i))  # i = week_row_index for grouping

            i = j
        else:
            i += 1

    return assignments, rows


def resolve_dates(assignments_raw, year=2026, month=6):
    """Convert day numbers to actual date objects. Handle month transitions."""
    resolved = []
    for (physician, day_num, shift_code, week_idx) in assignments_raw:
        if day_num is None:
            continue
        # June 2026: days 1-30 are June, if day_num > 28 could be late month, if day_num < 5 could be July
        # We need to handle the week that spans month boundary
        # Strategy: track month based on day number sequence
        # For June 2026 schedule, most days are June (1-30)
        # Days <= 30 are June, days that appear after day 28 but are small (1-7) are July
        # We'll handle this by checking context - for now use simple heuristic
        try:
            d = date(year, month, int(day_num))
            resolved.append((physician, d, shift_code))
        except ValueError:
            # Day out of range - might be July
            try:
                d = date(year, month + 1, int(day_num))
                resolved.append((physician, d, shift_code))
            except ValueError:
                pass
    return resolved


def parse_file(filepath):
    """Full parse returning list of (physician, date, shift_code)."""
    print(f"\nParsing: {filepath}")
    assignments_raw, rows = parse_schedule(filepath)
    resolved = resolve_dates(assignments_raw)
    print(f"  Raw assignments found: {len(assignments_raw)}")
    print(f"  Resolved assignments: {len(resolved)}")
    return resolved


# ---- ALTERNATIVE PARSER: more robust row-by-row ----

def parse_file_v2(filepath, year=2026, month=6):
    """Robust parser that handles the schedule format carefully."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)

    assignments = []  # (physician, date_obj, shift_code)

    i = 0
    current_dates = {}  # col_idx (0-6) -> date_obj

    while i < len(rows):
        row = rows[i]

        # Ensure row has enough columns
        row = list(row) + [None] * 10  # pad

        col_a = str(row[0]).strip() if row[0] is not None else ''
        col_b = str(row[1]).strip() if row[1] is not None else ''

        # Header row detection: col B == 'SUN'
        if col_b == 'SUN':
            # Next row: date numbers
            i += 1
            if i < len(rows):
                date_row = list(rows[i]) + [None] * 10
                current_dates = {}
                for col_idx in range(7):  # 0=SUN ... 6=SAT
                    val = date_row[col_idx + 1]  # cols B-H
                    if val is not None:
                        try:
                            day_num = int(val)
                            # Resolve to actual date
                            # Heuristic: if day_num > 25 and we're tracking a month boundary
                            try:
                                d = date(year, month, day_num)
                            except ValueError:
                                try:
                                    d = date(year, month + 1, day_num) if month < 12 else date(year + 1, 1, day_num)
                                except ValueError:
                                    d = None
                            current_dates[col_idx] = d
                        except (ValueError, TypeError):
                            pass
            i += 1
            continue

        # Site row detection: col A has a known site label
        site_label = col_a
        if site_label in SITE_MAP or site_label in SKIP_SITES:
            # Next row should be time row
            if i + 1 < len(rows):
                time_row = list(rows[i + 1]) + [None] * 10
                time_val_raw = str(time_row[0]).strip() if time_row[0] is not None else ''

                if site_label in SKIP_SITES:
                    i += 2
                    continue

                site_code = SITE_MAP.get(site_label, site_label)
                time_code = TIME_MAP.get(time_val_raw) or TIME_MAP.get('  ' + time_val_raw)
                if time_code is None:
                    # Try stripping
                    stripped = time_val_raw.strip()
                    time_code = TIME_MAP.get(stripped)

                if time_code and current_dates:
                    shift_code = f"{time_code} {site_code}"
                    # Extract physicians from cols B-H (indices 1-7)
                    for col_idx in range(7):
                        physician = row[col_idx + 1]
                        if physician is not None:
                            physician = str(physician).strip()
                            if physician and physician.lower() not in ('none', ''):
                                d = current_dates.get(col_idx)
                                if d:
                                    assignments.append((physician, d, shift_code))
                i += 2
                continue

        i += 1

    return assignments


# ---- MAIN ----

file_machine = r"C:\Users\kskob\Dropbox\KEAclaude\KEAsked\Request-Imports\test-schedule.xlsx"
file_human   = r"C:\Users\kskob\Dropbox\KEAclaude\KEAsked\Request-Imports\June26-complete.xlsx"

print("=" * 70)
print("PARSING MACHINE-GENERATED SCHEDULE")
print("=" * 70)
machine_assignments = parse_file_v2(file_machine)
print(f"Total assignments parsed: {len(machine_assignments)}")

print("\n" + "=" * 70)
print("PARSING HUMAN-GENERATED SCHEDULE")
print("=" * 70)
human_assignments = parse_file_v2(file_human)
print(f"Total assignments parsed: {len(human_assignments)}")

# Show sample from each
print("\n--- Sample machine assignments (first 10) ---")
for a in sorted(machine_assignments, key=lambda x: x[1])[:10]:
    print(f"  {a[0]:20s} {a[1]} {a[2]}")

print("\n--- Sample human assignments (first 10) ---")
for a in sorted(human_assignments, key=lambda x: x[1])[:10]:
    print(f"  {a[0]:20s} {a[1]} {a[2]}")

# Build lookup structures
def build_physician_schedule(assignments):
    """Returns dict: physician -> list of (date, shift_code)"""
    sched = defaultdict(list)
    for (phys, d, sc) in assignments:
        sched[phys].append((d, sc))
    for phys in sched:
        sched[phys].sort(key=lambda x: x[0])
    return sched

machine_sched = build_physician_schedule(machine_assignments)
human_sched   = build_physician_schedule(human_assignments)

# ============================================================
print("\n" + "=" * 70)
print("ANALYSIS 1: Same shift on consecutive days (machine schedule)")
print("=" * 70)

violations = []
for phys, shifts in machine_sched.items():
    for idx in range(len(shifts) - 1):
        d1, sc1 = shifts[idx]
        d2, sc2 = shifts[idx + 1]
        if sc1 == sc2 and (d2 - d1).days == 1:
            violations.append((phys, d1, sc1, d2))

if violations:
    print(f"Found {len(violations)} consecutive-day same-shift violations:")
    for (phys, d1, sc, d2) in sorted(violations):
        print(f"  {phys:20s}  {d1} -> {d2}  [{sc}]")
else:
    print("No consecutive-day same-shift violations found.")

# ============================================================
print("\n" + "=" * 70)
print("ANALYSIS 2: Singleton 2400h shifts (machine schedule)")
print("=" * 70)

singletons_2400 = []
for phys, shifts in machine_sched.items():
    # Get all dates with 2400h
    night_dates = sorted([d for (d, sc) in shifts if '2400h' in sc])
    for idx, nd in enumerate(night_dates):
        prev_date = night_dates[idx - 1] if idx > 0 else None
        next_date = night_dates[idx + 1] if idx < len(night_dates) - 1 else None
        has_prev = prev_date is not None and (nd - prev_date).days == 1
        has_next = next_date is not None and (next_date - nd).days == 1
        if not has_prev and not has_next:
            singletons_2400.append((phys, nd))

print(f"Singleton 2400h shifts: {len(singletons_2400)}")
for (phys, nd) in sorted(singletons_2400):
    print(f"  {phys:20s}  {nd}")

# ============================================================
print("\n" + "=" * 70)
print("ANALYSIS 3: Fill rate comparison")
print("=" * 70)

EXPECTED_TOTAL = 630  # 21 shifts/day × 30 days

# Count filled shifts (non-empty physician assignments)
machine_filled = len(machine_assignments)
human_filled   = len(human_assignments)

# Count unique (date, shift_code) slots filled
machine_slots = set((d, sc) for (p, d, sc) in machine_assignments)
human_slots   = set((d, sc) for (p, d, sc) in human_assignments)

print(f"Machine schedule:")
print(f"  Total physician assignments: {machine_filled}")
print(f"  Unique (date, shift_code) slots filled: {len(machine_slots)}")
print(f"  Expected total slots: {EXPECTED_TOTAL}")
print(f"  Apparent unfilled slots: {EXPECTED_TOTAL - len(machine_slots)} (based on expected)")

print(f"\nHuman schedule:")
print(f"  Total physician assignments: {human_filled}")
print(f"  Unique (date, shift_code) slots filled: {len(human_slots)}")
print(f"  Apparent unfilled slots: {EXPECTED_TOTAL - len(human_slots)}")

# ============================================================
print("\n" + "=" * 70)
print("ANALYSIS 4: Per-physician shift counts")
print("=" * 70)

all_physicians = sorted(set(list(machine_sched.keys()) + list(human_sched.keys())))

print(f"{'Physician':<25} {'Machine':>8} {'Human':>8} {'Diff':>8}")
print("-" * 55)
for phys in all_physicians:
    mc = len(machine_sched.get(phys, []))
    hc = len(human_sched.get(phys, []))
    diff = mc - hc
    flag = " <-- LARGE DIFF" if abs(diff) >= 3 else ""
    print(f"  {phys:<23} {mc:>8} {hc:>8} {diff:>+8}{flag}")

# ============================================================
print("\n" + "=" * 70)
print("ANALYSIS 5: Physicians in human schedule not in machine schedule")
print("=" * 70)

human_only = [p for p in human_sched if p not in machine_sched or len(machine_sched[p]) == 0]
machine_only = [p for p in machine_sched if p not in human_sched or len(human_sched[p]) == 0]
few_in_machine = [p for p in human_sched if p in machine_sched and 0 < len(machine_sched[p]) <= 2 and len(human_sched[p]) >= 3]

print("In human but NOT in machine schedule:")
for p in sorted(human_only):
    print(f"  {p} (human: {len(human_sched[p])} shifts)")

print("\nIn machine but NOT in human schedule:")
for p in sorted(machine_only):
    print(f"  {p} (machine: {len(machine_sched[p])} shifts)")

print("\nIn human with >= 3 shifts, but very few (<=2) in machine:")
for p in sorted(few_in_machine):
    print(f"  {p} (machine: {len(machine_sched[p])}, human: {len(human_sched[p])})")

# ============================================================
print("\n" + "=" * 70)
print("ANALYSIS 6: Shift time distribution (machine vs human)")
print("=" * 70)

TIME_SLOTS = ['0600h', '0900h', '1000h', '1200h', '1400h', '1500h', '1600h', '1700h', '1800h', '2000h', '2400h']

machine_time_counts = defaultdict(int)
human_time_counts   = defaultdict(int)

for (p, d, sc) in machine_assignments:
    for ts in TIME_SLOTS:
        if sc.startswith(ts):
            machine_time_counts[ts] += 1

for (p, d, sc) in human_assignments:
    for ts in TIME_SLOTS:
        if sc.startswith(ts):
            human_time_counts[ts] += 1

print(f"{'Time Slot':<12} {'Machine':>8} {'Human':>8} {'Diff':>8}")
print("-" * 42)
for ts in TIME_SLOTS:
    mc = machine_time_counts[ts]
    hc = human_time_counts[ts]
    diff = mc - hc
    print(f"  {ts:<10} {mc:>8} {hc:>8} {diff:>+8}")

# Also show site distribution
print()
SITES = ['RAH A side', 'RAH B side', 'NEHC', 'RAH I side', 'RAH F side']
machine_site_counts = defaultdict(int)
human_site_counts   = defaultdict(int)

for (p, d, sc) in machine_assignments:
    for site in SITES:
        if site in sc:
            machine_site_counts[site] += 1

for (p, d, sc) in human_assignments:
    for site in SITES:
        if site in sc:
            human_site_counts[site] += 1

print(f"{'Site':<15} {'Machine':>8} {'Human':>8} {'Diff':>8}")
print("-" * 45)
for site in SITES:
    mc = machine_site_counts[site]
    hc = human_site_counts[site]
    diff = mc - hc
    print(f"  {site:<13} {mc:>8} {hc:>8} {diff:>+8}")

# ============================================================
print("\n" + "=" * 70)
print("ANALYSIS 7: 2400h run analysis")
print("=" * 70)

def get_runs(night_dates):
    """Returns list of runs (each run is a list of consecutive dates)."""
    if not night_dates:
        return []
    runs = []
    current_run = [night_dates[0]]
    for nd in night_dates[1:]:
        if (nd - current_run[-1]).days == 1:
            current_run.append(nd)
        else:
            runs.append(current_run)
            current_run = [nd]
    runs.append(current_run)
    return runs

print("\n--- Machine schedule 2400h runs ---")
for phys in sorted(machine_sched.keys()):
    night_dates = sorted([d for (d, sc) in machine_sched[phys] if '2400h' in sc])
    if not night_dates:
        continue
    runs = get_runs(night_dates)
    run_desc = ', '.join(
        f"{r[0].strftime('%b%d')}" if len(r) == 1 else f"{r[0].strftime('%b%d')}-{r[-1].strftime('%b%d')}({len(r)}d)"
        for r in runs
    )
    print(f"  {phys:<22} total={len(night_dates):2d}  runs: {run_desc}")

print("\n--- Human schedule 2400h runs ---")
for phys in sorted(human_sched.keys()):
    night_dates = sorted([d for (d, sc) in human_sched[phys] if '2400h' in sc])
    if not night_dates:
        continue
    runs = get_runs(night_dates)
    run_desc = ', '.join(
        f"{r[0].strftime('%b%d')}" if len(r) == 1 else f"{r[0].strftime('%b%d')}-{r[-1].strftime('%b%d')}({len(r)}d)"
        for r in runs
    )
    print(f"  {phys:<22} total={len(night_dates):2d}  runs: {run_desc}")

# ============================================================
print("\n" + "=" * 70)
print("ANALYSIS 8: Rule deviations and anomalies")
print("=" * 70)

# 8a: Physicians working multiple shifts on same day
print("\n8a. Physicians with multiple shifts on same day (machine):")
multi_day = []
for phys, shifts in machine_sched.items():
    by_date = defaultdict(list)
    for (d, sc) in shifts:
        by_date[d].append(sc)
    for d, scs in by_date.items():
        if len(scs) > 1:
            multi_day.append((phys, d, scs))
if multi_day:
    for (phys, d, scs) in sorted(multi_day):
        print(f"  {phys:20s} {d}: {scs}")
else:
    print("  None found.")

# 8b: Long runs of consecutive work days (any shift)
print("\n8b. Physicians with 4+ consecutive work days (machine):")
for phys, shifts in sorted(machine_sched.items()):
    work_dates = sorted(set(d for (d, sc) in shifts))
    if len(work_dates) < 4:
        continue
    runs = []
    current = [work_dates[0]]
    for d in work_dates[1:]:
        if (d - current[-1]).days == 1:
            current.append(d)
        else:
            runs.append(current)
            current = [d]
    runs.append(current)
    long_runs = [r for r in runs if len(r) >= 4]
    if long_runs:
        for r in long_runs:
            print(f"  {phys:20s} {r[0]} to {r[-1]} ({len(r)} days)")

# 8c: Consecutive days comparison - how many in machine vs human
machine_consec = sum(1 for phys, shifts in machine_sched.items()
                     for idx in range(len(shifts)-1)
                     if (shifts[idx+1][0] - shifts[idx][0]).days == 1)
human_consec   = sum(1 for phys, shifts in human_sched.items()
                     for idx in range(len(shifts)-1)
                     if (shifts[idx+1][0] - shifts[idx][0]).days == 1)

print(f"\n8c. Total consecutive-day pairs: Machine={machine_consec}, Human={human_consec}")

# 8d: Shift variety per physician
print("\n8d. Physicians assigned many different shift types (machine):")
for phys in sorted(machine_sched.keys()):
    shift_types = set(sc for (d, sc) in machine_sched[phys])
    if len(shift_types) > 3:
        print(f"  {phys:20s} {len(shift_types)} distinct shift types: {sorted(shift_types)}")

# 8e: Days with 0 coverage in machine
print("\n8e. Shift slots with NO physician assigned (machine):")
# Get all shift_codes in human schedule to know what slots should exist
all_shift_codes = set(sc for (p, d, sc) in human_assignments)
all_dates_machine = set(d for (p, d, sc) in machine_assignments)
machine_slot_set = set((d, sc) for (p, d, sc) in machine_assignments)

uncovered = []
for d in sorted(all_dates_machine):
    for sc in sorted(all_shift_codes):
        if (d, sc) not in machine_slot_set:
            uncovered.append((d, sc))

if uncovered:
    print(f"  {len(uncovered)} unfilled (date, shift_code) pairs in machine schedule:")
    # Group by shift code
    by_sc = defaultdict(list)
    for (d, sc) in uncovered:
        by_sc[sc].append(d)
    for sc in sorted(by_sc.keys()):
        dates = sorted(by_sc[sc])
        print(f"    {sc}: {len(dates)} days missing -- {[str(d) for d in dates[:5]]}{'...' if len(dates)>5 else ''}")
else:
    print("  No completely unfilled slots detected.")

# 8f: Overall summary statistics
print("\n8f. Summary statistics:")
print(f"  Machine: {len(machine_sched)} physicians, {len(machine_assignments)} total assignments")
print(f"  Human:   {len(human_sched)} physicians, {len(human_assignments)} total assignments")

# Physicians shared between both
shared = set(machine_sched.keys()) & set(human_sched.keys())
print(f"  Physicians in both: {len(shared)}")
print(f"  Only in machine: {sorted(set(machine_sched.keys()) - set(human_sched.keys()))}")
print(f"  Only in human: {sorted(set(human_sched.keys()) - set(machine_sched.keys()))}")

print("\n" + "=" * 70)
print("ANALYSIS COMPLETE")
print("=" * 70)

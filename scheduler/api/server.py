"""
FastAPI server for the physician scheduling system.

Launched as a subprocess by the Electron main process.
Listens on 127.0.0.1:5000.

Usage (standalone):
    python -m scheduler.api.server
"""

from __future__ import annotations

import asyncio
import calendar
import datetime
import io
import traceback
from pathlib import Path
from typing import Any

import yaml
import uvicorn
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from scheduler.api.schemas import (
    AssignmentSchema,
    CandidateSchema,
    CandidatesResponse,
    DetectFlatResponse,
    GenerateCachedRequest,
    GenerateRequest,
    ImportFlatRequest,
    ImportRequest,
    ManualAssignRequest,
    ManualAssignResponse,
    PhysicianImportResult,
    OnCallAssignmentSchema,
    PhysicianInfo,
    PhysiciansResponse,
    ImportDirectoryResponse,
    LoadScheduleRequest,
    ScheduleResponse,
    ScheduleStatsSchema,
    ShiftSchema,
    UnfilledSlotSchema,
    ValidationIssueSchema,
    ViolationSchema,
)
import os

from scheduler.backend.config import load_roster
from scheduler.backend.generator import (
    Assignment,
    OnCallAssignment,
    ScheduleGenerator,
    ScheduleResult,
    ScheduleStats,
    UnfilledSlot,
    _HARD_VIOLATION_RULES,
    generate_schedule,
)
try:
    from scheduler.backend.generator_cpsat import CpsatScheduleGenerator
    _CPSAT_AVAILABLE = True
except Exception:  # pragma: no cover
    _CPSAT_AVAILABLE = False
from scheduler.backend.importer import import_directory, import_single_file
from scheduler.backend.importer_flat import import_flat_file
from scheduler.backend.models import PhysicianSubmission
from scheduler.backend.shifts import ALL_SHIFT_CODES, BLOCKS, SHIFT_TO_BLOCK, Shift


def _v(v) -> ViolationSchema:
    """Convert a ViolationReason to ViolationSchema, including is_hard flag."""
    return ViolationSchema(
        rule=v.rule,
        description=v.description,
        is_hard=v.rule in _HARD_VIOLATION_RULES,
    )
from scheduler.backend.validator import validate

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(title="Physician Scheduler API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],    # Electron renderer connects from file:// or localhost
    allow_methods=["*"],
    allow_headers=["*"],
)

import os as _os
import sys as _sys

if _os.environ.get("CONFIG_DIR"):
    _CONFIG_DIR = Path(_os.environ["CONFIG_DIR"])
elif getattr(_sys, "frozen", False):
    # PyInstaller bundle — config lives next to the executable
    _CONFIG_DIR = Path(_sys.executable).parent / "config"
else:
    _CONFIG_DIR = Path(__file__).parent.parent / "config"

_SCHEDULER_CONFIG_PATH = _CONFIG_DIR / "scheduler_config.yaml"

# ---------------------------------------------------------------------------
# In-memory state (single-user desktop app — no database needed)
# ---------------------------------------------------------------------------

_state: dict[str, Any] = {
    "submissions": [],
    "roster": {},
    "scheduler_config": {},
    "generator": None,
    "result": None,
    "year": None,
    "month": None,
    "directory": None,
    "source_file": None,
    "progress": {"current": 0, "total": 0, "running": False, "best_unfilled": None},
}


def _apply_roster(submissions: list[PhysicianSubmission], roster: dict) -> None:
    """
    Merge per-physician config (rule overrides etc.) into each submission.
    Matching is case-insensitive so flat-file names like "BRAUN" match
    roster IDs like "Braun".
    """
    lower_roster = {k.lower(): v for k, v in roster.items()}
    for sub in submissions:
        cfg = roster.get(sub.physician_id) or lower_roster.get(sub.physician_id.lower())
        if cfg:
            sub.rule_overrides = dict(cfg.rule_overrides)


def _build_import_results(
    submissions: list[PhysicianSubmission],
) -> list[PhysicianImportResult]:
    results = []
    for sub in submissions:
        vr = validate(sub)
        valid_days = sum(1 for d in sub.days if d.is_valid_day)
        valid_blocks = sum(len(d.available_blocks) for d in sub.days if d.is_valid_day)
        valid_weekends = sum(1 for d in sub.days if d.is_valid_weekend)
        anchored = sum(1 for d in sub.days if d.is_anchored)
        results.append(PhysicianImportResult(
            physician_id=sub.physician_id,
            physician_name=sub.physician_name,
            shifts_requested=sub.shifts_requested,
            shifts_min=sub.shifts_min,
            shifts_max=sub.shifts_max,
            shifts_2400h_requested=sub.shifts_2400h_requested,
            shifts_0600h_requested=sub.shifts_0600h_requested,
            valid_days=valid_days,
            valid_blocks=valid_blocks,
            valid_weekend_days=valid_weekends,
            anchored_days=anchored,
            issues=[
                ValidationIssueSchema(
                    severity=i.severity,
                    rule=i.rule,
                    message=i.message,
                    physician_id=i.physician_id or sub.physician_id,
                )
                for i in vr.issues
            ],
            is_valid=vr.is_valid,
        ))
    return results


def _load_scheduler_config() -> dict:
    import os
    with _SCHEDULER_CONFIG_PATH.open(encoding="utf-8") as fh:
        cfg = yaml.safe_load(fh)
    return cfg


def _shift_to_schema(shift: Shift) -> ShiftSchema:
    return ShiftSchema(
        time=shift.time,
        site=shift.site,
        code=shift.code,
        site_group=shift.site_group.value,
    )


def _result_to_response(result: ScheduleResult) -> ScheduleResponse:
    return ScheduleResponse(
        year=result.year,
        month=result.month,
        assignments=[
            AssignmentSchema(
                date=a.date.isoformat(),
                shift=_shift_to_schema(a.shift),
                physician_id=a.physician_id,
                physician_name=a.physician_name,
                is_manual=a.is_manual,
            )
            for a in result.assignments
        ],
        unfilled=[
            UnfilledSlotSchema(
                date=u.date.isoformat(),
                shift=_shift_to_schema(u.shift),
                candidates=[
                    CandidateSchema(
                        physician_id=c.physician_id,
                        physician_name=c.physician_name,
                        violations=[
                            _v(v)
                            for v in c.violations
                        ],
                        is_hard_blocked=c.is_hard_blocked,
                    )
                    for c in u.candidates
                ],
            )
            for u in result.unfilled
        ],
        issues=result.issues,
        stats=ScheduleStatsSchema(**result.stats.__dict__) if result.stats else None,
        on_calls=[
            OnCallAssignmentSchema(
                date=oc.date.isoformat(),
                call_type=oc.call_type,
                physician_id=oc.physician_id,
                physician_name=oc.physician_name,
            )
            for oc in result.on_calls
        ],
    )


# ---------------------------------------------------------------------------
# Generator rebuild helper
# ---------------------------------------------------------------------------

def _require_generator() -> tuple:
    """
    Return (gen, result) from _state, rebuilding gen from cached submissions
    if it was lost (e.g. after loading a schedule from file or server restart).

    Raises HTTPException(400) if result or submissions are absent.
    """
    result: ScheduleResult | None = _state.get("result")
    if result is None:
        raise HTTPException(status_code=400, detail="No schedule in memory.")

    gen: ScheduleGenerator | None = _state.get("generator")
    if gen is not None:
        return gen, result

    # Generator absent — try to rebuild from cached submissions.
    submissions: list[PhysicianSubmission] = _state.get("submissions") or []
    if not submissions:
        raise HTTPException(
            status_code=400,
            detail="No schedule in memory. Re-import physician preferences to enable this action.",
        )
    roster = _state.get("roster") or {}
    cfg = _state.get("scheduler_config") or {}
    if _CPSAT_AVAILABLE:
        gen = CpsatScheduleGenerator(submissions, roster, cfg)
    else:
        gen = ScheduleGenerator(submissions, roster, cfg)
    for a in result.assignments:
        gen._assign(a.physician_id, a.date, a.shift)
    _state["generator"] = gen   # cache for subsequent calls
    return gen, result


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/api/detect-flat", response_model=DetectFlatResponse)
def detect_flat(file: str) -> DetectFlatResponse:
    """
    Read the first data row of a flat file and return the year/month found.
    Used by the frontend to auto-fill the month/year selectors.
    """
    import openpyxl
    fp = Path(file)
    if not fp.is_file():
        raise HTTPException(status_code=400, detail=f"File not found: {file}")
    try:
        from collections import Counter
        from scheduler.backend.importer_flat import _to_date
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        ws = wb.active
        counts: Counter = Counter()
        for row in ws.iter_rows(min_row=2, max_row=200, values_only=True):
            if not row[0] or not row[1]:
                continue
            d = _to_date(row[1])
            if d:
                counts[(d.year, d.month)] += 1
        wb.close()
        if counts:
            (year, month), _ = counts.most_common(1)[0]
            return DetectFlatResponse(year=year, month=month)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    raise HTTPException(status_code=422, detail="Could not detect year/month from file.")


@app.get("/api/physicians", response_model=PhysiciansResponse)
def get_physicians() -> PhysiciansResponse:
    """Return all physicians from the roster."""
    roster = load_roster()
    return PhysiciansResponse(
        physicians=[
            PhysicianInfo(
                id=cfg.id,
                name=cfg.name,
                active=cfg.active,
                max_consecutive_shifts=cfg.max_consecutive_shifts,
                group_b_site_preference=cfg.group_b_site_preference,
                forbidden_sites=cfg.forbidden_sites,
            )
            for cfg in roster.values()
        ]
    )


@app.post("/api/import", response_model=ImportDirectoryResponse)
def import_submissions(body: ImportRequest) -> ImportDirectoryResponse:
    """Import all .xlsx files from a directory for the given year/month."""
    directory = Path(body.directory)
    if not directory.is_dir():
        raise HTTPException(status_code=400, detail=f"Directory not found: {body.directory}")
    try:
        roster = load_roster()
        scheduler_cfg = _load_scheduler_config()
        submissions = import_directory(directory, body.year, body.month)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))

    _apply_roster(submissions, roster)
    results = _build_import_results(submissions)
    _state.update(submissions=submissions, roster=roster, scheduler_config=scheduler_cfg,
                  year=body.year, month=body.month, directory=body.directory, source_file=None)

    valid_count = sum(1 for r in results if r.is_valid)
    return ImportDirectoryResponse(year=body.year, month=body.month, directory=body.directory,
                                   physicians=results, total_physicians=len(results),
                                   valid_physicians=valid_count)


@app.post("/api/import-flat", response_model=ImportDirectoryResponse)
def import_flat(body: ImportFlatRequest) -> ImportDirectoryResponse:
    """Import a single flat-table .xlsx file (all physicians in one sheet)."""
    file_path = Path(body.file)
    if not file_path.is_file():
        raise HTTPException(status_code=400, detail=f"File not found: {body.file}")
    try:
        roster = load_roster()
        scheduler_cfg = _load_scheduler_config()
        submissions = import_flat_file(file_path, body.year, body.month)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))

    _apply_roster(submissions, roster)
    results = _build_import_results(submissions)
    _state.update(submissions=submissions, roster=roster, scheduler_config=scheduler_cfg,
                  year=body.year, month=body.month, directory=None, source_file=str(file_path))

    valid_count = sum(1 for r in results if r.is_valid)
    return ImportDirectoryResponse(year=body.year, month=body.month,
                                   directory=str(file_path),
                                   physicians=results, total_physicians=len(results),
                                   valid_physicians=valid_count)


@app.get("/api/generate-progress")
def get_generate_progress() -> dict:
    return _state["progress"]


@app.post("/api/generate", response_model=ScheduleResponse)
async def generate(body: GenerateCachedRequest) -> ScheduleResponse:
    """
    Generate a schedule from previously imported submissions (runs 300 iterations).
    Call /api/import or /api/import-flat first.
    Poll /api/generate-progress for live iteration count.
    """
    submissions: list[PhysicianSubmission] = _state["submissions"]
    if not submissions or _state["year"] != body.year or _state["month"] != body.month:
        raise HTTPException(
            status_code=400,
            detail=f"No imported submissions for {body.year}-{body.month:02d}. "
                   "Call /api/import or /api/import-flat first.",
        )

    roster = _state["roster"]
    cfg = _state["scheduler_config"]
    use_cpsat = _CPSAT_AVAILABLE
    n_iterations = 400
    cpsat_time_limit = 600.0

    if use_cpsat:
        # CP-SAT: indeterminate progress — show 50% "Solving…" until done.
        _state["progress"] = {"current": 50, "total": 100, "running": True, "best_unfilled": None, "solver": "cpsat", "time_limit": int(cpsat_time_limit)}
    else:
        _state["progress"] = {"current": 0, "total": n_iterations, "running": True, "best_unfilled": None, "solver": "greedy"}

    def progress_cb(current: int, total: int, best_score: float) -> None:
        _state["progress"]["current"] = current
        # Derive approximate unfilled count from score: score = -unfilled*1000 + ...
        # Just show the raw best score for now
        _state["progress"]["best_unfilled"] = round(-best_score / 1000)

    try:
        if use_cpsat:
            gen = CpsatScheduleGenerator(submissions, roster, cfg)
            result = await asyncio.to_thread(
                gen.generate, body.year, body.month, cpsat_time_limit, 8, progress_cb
            )
        else:
            gen = ScheduleGenerator(submissions, roster, cfg)
            result = await asyncio.to_thread(
                gen.run_best_of, n_iterations, body.year, body.month, progress_cb
            )
        # Post-solve repair: juggle adjacent assignments to fill remaining gaps
        if result.unfilled:
            result = await asyncio.to_thread(gen.repair_pass, result, 50)
        # Sync issues list: only keep entries for slots that remain unfilled
        # (repair pass and Claude may have filled slots that still appear in issues)
        unfilled_keys = {
            f"{u.date.strftime('%b %d')} {u.shift.code}" for u in result.unfilled
        }
        result.issues = [i for i in result.issues if any(k in i for k in unfilled_keys)] \
            if result.unfilled else []
        # Assign on-call shifts after the regular schedule is complete
        result = await asyncio.to_thread(gen.assign_on_calls, result)
    except Exception as exc:
        _state["progress"]["running"] = False
        raise HTTPException(status_code=500, detail=f"Generation failed: {exc}\n{traceback.format_exc()}")

    final_total = 100 if use_cpsat else n_iterations
    _state["progress"] = {"current": final_total, "total": final_total, "running": False, "best_unfilled": len(result.unfilled)}
    _state["generator"] = gen
    _state["result"] = result

    return _result_to_response(result)


@app.get("/api/schedule", response_model=ScheduleResponse)
def get_schedule() -> ScheduleResponse:
    """Return the most recently generated schedule."""
    result: ScheduleResult | None = _state.get("result")
    if result is None:
        raise HTTPException(status_code=404, detail="No schedule generated yet.")
    return _result_to_response(result)




@app.get("/api/export")
def export_schedule() -> StreamingResponse:
    """Export the current schedule as an Excel file in the same layout as the human schedule."""
    result: ScheduleResult | None = _state.get("result")
    if result is None:
        raise HTTPException(status_code=404, detail="No schedule generated yet.")

    wb = _build_export_workbook(result)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"schedule_{result.year}_{result.month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Shift rows in display order: (site_label, time_label, time_code, site_code)
# time_code/site_code == None means it is an on-call row filled from result.on_calls.
_EXPORT_SHIFTS = [
    ("DOC",       "Day On Call", None,    None),           # Day on call row
    ("RAH A",     "0600-1200",  "0600h",  "RAH A side"),
    ("RAH B",     "0600-1200",  "0600h",  "RAH B side"),
    ("NECHC",     "0600-1400",  "0600h",  "NEHC"),
    ("RAH I",     "0600-1400",  "0600h",  "RAH I side"),
    ("NECHC",     "0900-1700",  "0900h",  "NEHC"),
    ("RAH I",     "1000-1800",  "1000h",  "RAH I side"),
    ("RAH A",     "1200-1800",  "1200h",  "RAH A side"),
    ("RAH B",     "1200-1800",  "1200h",  "RAH B side"),
    ("NECHC",     "1200-2000",  "1200h",  "NEHC"),
    ("RAH I",     "1400-2200",  "1400h",  "RAH I side"),
    ("NECHC",     "1500-2300",  "1500h",  "NEHC"),
    ("NOC",       "Night On Call", None,  None),           # Night on call row
    ("RAH Float", "1600-0459",  "1600h",  "RAH F side"),
    ("NECHC",     "1700-0100",  "1700h",  "NEHC"),
    ("RAH A",     "1800-0000",  "1800h",  "RAH A side"),
    ("RAH B",     "1800-0000",  "1800h",  "RAH B side"),
    ("RAH I",     "1800-0200",  "1800h",  "RAH I side"),
    ("NECHC",     "2000-0400",  "2000h",  "NEHC"),
    ("RAH A",     "2400-0600",  "2400h",  "RAH A side"),
    ("RAH B",     "2400-0600",  "2400h",  "RAH B side"),
    ("NECHC",     "2400-0800",  "2400h",  "NEHC"),
    ("RAH I",     "2400-0800",  "2400h",  "RAH I side"),
]

_HDR_FILL  = PatternFill("solid", fgColor="1E293B")
_HDR_FONT  = Font(bold=True, color="FFFFFF", size=9)
_DATE_FILL = PatternFill("solid", fgColor="334155")
_DATE_FONT = Font(bold=True, color="FFFFFF", size=9)
_LABEL_FONT = Font(bold=True, size=9)
_TIME_FONT  = Font(italic=True, color="64748B", size=8)
_CELL_FONT  = Font(size=9)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT   = Alignment(horizontal="left",   vertical="center")
_THIN   = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_NIGHT_FILL = PatternFill("solid", fgColor="EFF6FF")   # light blue for 2400h rows
_AM_FILL    = PatternFill("solid", fgColor="F0FDF4")   # light green for 0600h rows


def _build_export_workbook(result: ScheduleResult) -> openpyxl.Workbook:
    # Index regular assignments by (date, shift_code) -> physician_name
    index: dict[tuple, str] = {
        (a.date, a.shift.code): a.physician_name
        for a in result.assignments
    }
    # Index on-call assignments by (date, call_type) -> physician_name
    call_index: dict[tuple, str] = {
        (oc.date, oc.call_type): oc.physician_name
        for oc in result.on_calls
    }

    days_in_month = calendar.monthrange(result.year, result.month)[1]
    all_dates = [datetime.date(result.year, result.month, d) for d in range(1, days_in_month + 1)]

    # Group dates into Sun-starting weeks
    weeks: list[list[datetime.date | None]] = []
    # Find the first Sunday on or before the 1st
    first = all_dates[0]
    week_start = first - datetime.timedelta(days=(first.weekday() + 1) % 7)
    d = week_start
    while d <= all_dates[-1]:
        week = []
        for i in range(7):
            day = d + datetime.timedelta(days=i)
            week.append(day if day.month == result.month else None)
        weeks.append(week)
        d += datetime.timedelta(days=7)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{result.year}-{result.month:02d}"

    # Column widths: col A = label (12), cols B-H = day columns (11 each), col I = repeat label
    ws.column_dimensions["A"].width = 12
    for col_letter in ["B","C","D","E","F","G","H"]:
        ws.column_dimensions[col_letter].width = 11
    ws.column_dimensions["I"].width = 12

    row = 1
    day_names = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]

    for week_dates in weeks:
        # ── Week header: day names ──
        ws.row_dimensions[row].height = 14
        ws.cell(row, 1).fill = _HDR_FILL
        for col, name in enumerate(day_names, start=2):
            c = ws.cell(row, col, name)
            c.font = _HDR_FONT; c.fill = _HDR_FILL; c.alignment = _CENTER; c.border = _BORDER
        ws.cell(row, 9).fill = _HDR_FILL
        row += 1

        # ── Date numbers ──
        ws.row_dimensions[row].height = 13
        ws.cell(row, 1).fill = _DATE_FILL
        for col, d in enumerate(week_dates, start=2):
            c = ws.cell(row, col, d.day if d else "")
            c.font = _DATE_FONT; c.fill = _DATE_FILL; c.alignment = _CENTER; c.border = _BORDER
        ws.cell(row, 9).fill = _DATE_FILL
        row += 1

        # ── Shift rows (2 rows per shift: physician name + time range) ──
        for site_label, time_label, time_code, site_code in _EXPORT_SHIFTS:
            is_night  = time_code == "2400h"
            is_am     = time_code == "0600h"
            is_oncall = time_code is None   # DOC or NOC row

            # Row A: site label + physician names
            ws.row_dimensions[row].height = 14
            c = ws.cell(row, 1, site_label)
            c.font = _LABEL_FONT; c.alignment = _LEFT; c.border = _BORDER
            if is_night: c.fill = _NIGHT_FILL
            elif is_am:  c.fill = _AM_FILL

            for col, d in enumerate(week_dates, start=2):
                name = ""
                if d:
                    if is_oncall:
                        # DOC or NOC — look up from on-call index
                        name = call_index.get((d, site_label), "")
                    elif time_code and site_code:
                        shift_code = f"{time_code} {site_code}"
                        name = index.get((d, shift_code), "")
                c2 = ws.cell(row, col, name)
                c2.font = _CELL_FONT; c2.alignment = _CENTER; c2.border = _BORDER
                if is_night: c2.fill = _NIGHT_FILL
                elif is_am:  c2.fill = _AM_FILL

            # repeat label in col I
            c9 = ws.cell(row, 9, site_label)
            c9.font = _LABEL_FONT; c9.alignment = _LEFT; c9.border = _BORDER
            if is_night: c9.fill = _NIGHT_FILL
            elif is_am:  c9.fill = _AM_FILL
            row += 1

            # Row B: time range (greyed out)
            ws.row_dimensions[row].height = 11
            c = ws.cell(row, 1, time_label)
            c.font = _TIME_FONT; c.alignment = _LEFT; c.border = _BORDER
            for col in range(2, 9):
                ws.cell(row, col).border = _BORDER
            ws.cell(row, 9, time_label).font = _TIME_FONT
            row += 1

        # Small gap row between weeks
        row += 1

    return wb


# Reverse lookup: (site_label, time_label) -> (time_code, site_code)
_EXPORT_SHIFT_LOOKUP: dict[tuple[str, str], tuple] = {
    (site_label, time_label): (time_code, site_code)
    for site_label, time_label, time_code, site_code in _EXPORT_SHIFTS
}

# Flat shift-code -> Shift object lookup (used by xlsx loader)
_SHIFT_CODE_LOOKUP: dict[str, Shift] = {
    shift.code: shift
    for block in BLOCKS
    for shift in block
}


def _parse_schedule_xlsx(path: Path, roster: dict) -> ScheduleResult:
    """
    Parse a previously exported schedule xlsx back into a ScheduleResult.

    Reconstructs assignments, on-calls, unfilled slots, and stats.
    Physician IDs are resolved from the roster by display name.
    Fields not stored in the xlsx (solver_status, optimality_gap_pct,
    candidate lists) are set to None / [].
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # --- Year/month from sheet title (e.g. "2026-06") ---
    try:
        title = ws.title
        year, month = int(title[:4]), int(title[5:7])
    except Exception:
        raise ValueError(
            f"Cannot determine year/month from sheet title {ws.title!r}. "
            "Expected format: YYYY-MM."
        )

    # --- Name → ID reverse lookup (case-insensitive fallback) ---
    name_to_id: dict[str, str] = {}
    for pid, cfg in roster.items():
        name_to_id[cfg.name] = pid
        name_to_id[cfg.name.lower()] = pid

    def _resolve_id(name: str) -> str:
        return (
            name_to_id.get(name)
            or name_to_id.get(name.lower())
            or name
        )

    assignments: list[Assignment] = []
    on_calls: list[OnCallAssignment] = []
    unfilled: list[UnfilledSlot] = []

    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]

        # Detect week header: col B (index 1) == "SUN"
        if row[1] != "SUN":
            i += 1
            continue

        # Next row is the date number row
        i += 1
        if i >= len(rows):
            break
        date_row = rows[i]
        col_to_date: dict[int, datetime.date] = {}
        for c in range(1, 8):   # cols B–H → indices 1–7
            val = date_row[c]
            if val is not None and val != "":
                try:
                    col_to_date[c] = datetime.date(year, month, int(val))
                except (ValueError, TypeError):
                    pass

        # Parse shift pairs until gap row or next week header
        i += 1
        while i < len(rows):
            site_row = rows[i]

            # Gap row (col A is None/empty) → end of week
            if site_row[0] is None or str(site_row[0]).strip() == "":
                i += 1
                break

            # site_row is Row A of a shift pair; next row is Row B (time label)
            i += 1
            if i >= len(rows):
                break
            time_row = rows[i]
            i += 1

            site_label = str(site_row[0]).strip()
            time_label = str(time_row[0]).strip() if time_row[0] is not None else ""

            entry = _EXPORT_SHIFT_LOOKUP.get((site_label, time_label))
            if not entry:
                continue
            time_code, site_code = entry

            for c, d in col_to_date.items():
                cell_val = site_row[c]
                name = str(cell_val).strip() if cell_val is not None else ""

                if time_code is None:
                    # On-call row (DOC / NOC)
                    if name and name not in ("", "---", "None"):
                        on_calls.append(OnCallAssignment(
                            date=d,
                            call_type=site_label,
                            physician_id=_resolve_id(name),
                            physician_name=name,
                        ))
                else:
                    shift_code = f"{time_code} {site_code}"
                    shift = _SHIFT_CODE_LOOKUP.get(shift_code)
                    if shift is None:
                        continue
                    if name and name not in ("", "---", "None"):
                        assignments.append(Assignment(
                            date=d,
                            shift=shift,
                            physician_id=_resolve_id(name),
                            physician_name=name,
                        ))
                    else:
                        unfilled.append(UnfilledSlot(date=d, shift=shift, candidates=[]))

    # --- Compute stats from reconstructed assignments ---
    filled = len(assignments)
    total = filled + len(unfilled)
    group_a = sum(1 for a in assignments if a.shift.site_group.value == "A")
    group_b = filled - group_a

    physician_counts: dict[str, int] = {}
    for a in assignments:
        physician_counts[a.physician_id] = physician_counts.get(a.physician_id, 0) + 1

    # Singleton 2400h detection (isolated night = no adjacent night within 1 day)
    nights_by_pid: dict[str, list[datetime.date]] = {}
    for a in assignments:
        if a.shift.time == "2400h":
            nights_by_pid.setdefault(a.physician_id, []).append(a.date)
    physician_singletons: dict[str, int] = {}
    for pid, dates in nights_by_pid.items():
        dates_sorted = sorted(dates)
        count = sum(
            1 for j, d in enumerate(dates_sorted)
            if not (j > 0 and (d - dates_sorted[j - 1]).days == 1)
            and not (j < len(dates_sorted) - 1 and (dates_sorted[j + 1] - d).days == 1)
        )
        if count:
            physician_singletons[pid] = count

    stats = ScheduleStats(
        total_slots=total,
        filled_slots=filled,
        unfilled_slots=len(unfilled),
        group_a_count=group_a,
        group_b_count=group_b,
        group_a_pct=round(group_a / filled, 3) if filled else 0.0,
        group_b_pct=round(group_b / filled, 3) if filled else 0.0,
        physician_counts=physician_counts,
        physician_singletons=physician_singletons,
        solver_status=None,
        optimality_gap_pct=None,
    )

    return ScheduleResult(
        year=year,
        month=month,
        assignments=assignments,
        unfilled=unfilled,
        issues=[],
        stats=stats,
        on_calls=on_calls,
    )


@app.post("/api/load-schedule", response_model=ScheduleResponse)
def load_schedule(body: LoadScheduleRequest) -> ScheduleResponse:
    """
    Load a previously exported schedule xlsx and make it the active schedule.

    Manual assignment and swap endpoints require a generator (built during
    normal import+generate). When loading from file, those features show an
    error asking the user to re-import and regenerate.
    """
    fp = Path(body.file)
    if not fp.is_file():
        raise HTTPException(status_code=400, detail=f"File not found: {body.file}")
    if fp.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="File must be an .xlsx file.")

    roster = _state.get("roster") or {}
    if not roster:
        try:
            roster = load_roster()
            _state["roster"] = roster
        except Exception:
            roster = {}

    try:
        result = _parse_schedule_xlsx(fp, roster)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse schedule: {exc}")

    _state["result"] = result
    _state["year"] = result.year
    _state["month"] = result.month
    # Do not clear generator — if the user previously imported for this month,
    # manual assignment will still work. If months differ, the generator will
    # be wrong but the assign endpoint guards against that.

    return _result_to_response(result)


@app.post("/api/assign", response_model=ManualAssignResponse)
def manual_assign(body: ManualAssignRequest) -> ManualAssignResponse:
    """
    Manually assign a physician to a shift slot (human override).
    The assignment is force-applied even if rules are violated.
    Returns the list of rules that were broken for display.
    """
    gen, result = _require_generator()

    if body.shift_code not in ALL_SHIFT_CODES:
        raise HTTPException(status_code=400, detail=f"Unknown shift code: {body.shift_code!r}")

    if body.physician_id not in gen.submissions:
        raise HTTPException(
            status_code=400, detail=f"Physician {body.physician_id!r} not in current submissions."
        )

    try:
        d = datetime.date.fromisoformat(body.date)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date: {body.date!r}")

    # Find the Shift object
    shift_obj: Shift | None = None
    for block in BLOCKS:
        for s in block:
            if s.code == body.shift_code:
                shift_obj = s
                break
        if shift_obj:
            break

    # If slot is already occupied, unassign the previous physician first
    existing_assignment = next(
        (a for a in result.assignments if a.date == d and a.shift.code == body.shift_code),
        None,
    )
    if existing_assignment:
        gen._unassign(existing_assignment.physician_id, d, shift_obj)
        result.assignments = [
            a for a in result.assignments
            if not (a.date == d and a.shift.code == body.shift_code)
        ]

    violations = gen.assign_manual(body.physician_id, d, shift_obj)

    # Update the result: remove from unfilled if it was there, add to assignments
    sub = gen.submissions[body.physician_id]
    result.unfilled = [
        u for u in result.unfilled
        if not (u.date == d and u.shift.code == body.shift_code)
    ]
    result.assignments.append(
        Assignment(
            date=d,
            shift=shift_obj,
            physician_id=body.physician_id,
            physician_name=sub.physician_name,
            is_manual=True,
        )
    )

    return ManualAssignResponse(
        success=True,
        violations=[_v(v) for v in violations],
        message=(
            f"Assigned {sub.physician_name} to {body.shift_code} on {body.date}."
            + (f" {len(violations)} rule(s) overridden." if violations else " No rule violations.")
        ),
    )


@app.post("/api/check-violations")
def check_violations(body: ManualAssignRequest):
    """
    Check rule violations for assigning a physician to a slot WITHOUT assigning.
    Temporarily unassigns any current occupant for an accurate check, then restores.
    """
    gen, result = _require_generator()

    if body.shift_code not in ALL_SHIFT_CODES:
        raise HTTPException(status_code=400, detail=f"Unknown shift code: {body.shift_code!r}")

    if body.physician_id not in gen.submissions:
        raise HTTPException(
            status_code=400, detail=f"Physician {body.physician_id!r} not in submissions."
        )

    try:
        d = datetime.date.fromisoformat(body.date)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date: {body.date!r}")

    shift_obj: Shift | None = None
    for block in BLOCKS:
        for s in block:
            if s.code == body.shift_code:
                shift_obj = s
                break
        if shift_obj:
            break

    # Temporarily unassign current occupant for accurate check
    existing = next(
        (a for a in result.assignments if a.date == d and a.shift.code == body.shift_code),
        None,
    )
    if existing:
        gen._unassign(existing.physician_id, d, shift_obj)

    violations = gen._check_constraints(body.physician_id, d, shift_obj) or []

    # Restore the previous occupant
    if existing:
        gen._assign(existing.physician_id, d, shift_obj)

    return {
        "violations": [
            _v(v) for v in violations
        ]
    }


@app.get("/api/candidates", response_model=CandidatesResponse)
def get_candidates(date: str, shift_code: str) -> CandidatesResponse:
    """
    Recalculate fresh candidates for an unfilled slot.

    Unlike the candidates embedded in the schedule response (which are computed
    at generation time and can become stale after manual assignments), this
    endpoint uses the live generator state so it correctly reflects any
    assignments made since generation.

    Hard violations (consecutive_limit, spacing_23h, already_assigned_today,
    etc.) cause a physician to be excluded from the list entirely.
    Only soft violations are returned as warnings.
    """
    gen, result = _require_generator()

    if shift_code not in ALL_SHIFT_CODES:
        raise HTTPException(status_code=400, detail=f"Unknown shift code: {shift_code!r}")

    try:
        d = datetime.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date: {date!r}")

    shift_obj: Shift | None = None
    for block in BLOCKS:
        for s in block:
            if s.code == shift_code:
                shift_obj = s
                break
        if shift_obj:
            break

    if shift_obj is None:
        raise HTTPException(status_code=400, detail=f"Shift not found: {shift_code!r}")

    # Temporarily unassign current occupant so the check is accurate for
    # the "slot is empty" case — restoring afterwards.
    existing = next(
        (a for a in result.assignments if a.date == d and a.shift.code == shift_code),
        None,
    )
    if existing:
        gen._unassign(existing.physician_id, d, shift_obj)

    candidates = gen._near_miss_candidates(d, shift_obj, max_n=20)

    if existing:
        gen._assign(existing.physician_id, d, shift_obj)

    return CandidatesResponse(
        date=date,
        shift_code=shift_code,
        candidates=[
            CandidateSchema(
                physician_id=c.physician_id,
                physician_name=c.physician_name,
                violations=[
                    _v(v)
                    for v in c.violations
                ],
                is_hard_blocked=c.is_hard_blocked,
            )
            for c in candidates
        ],
    )


# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import logging
    import os
    log_path = Path(os.environ.get("CONFIG_DIR", ".")).parent / "scheduler_debug.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(str(log_path), encoding="utf-8"),
        ],
    )
    uvicorn.run(app, host="127.0.0.1", port=5000, log_level="warning", access_log=False)
